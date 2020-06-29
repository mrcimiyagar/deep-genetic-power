import copy
import os
import random
import subprocess

import numpy as np

from deap import base, creator, tools, algorithms
from scipy.stats import bernoulli
from bitstring import BitArray
import gc
import openpyxl

parallel = 1

dense_units = 0
lstm_units = 0
window_size = 0
lr = 0
ep = 0
bs = 0
act = 0
hact = 0
im = 0
opt = 0

if not os.path.exists('ModelsInfo.xlsx'):
    model_counter = 1
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    my_sheet.title = 'ModelsInfo'
    my_sheet.cell(row=model_counter, column=1).value = 'window'
    my_sheet.cell(row=model_counter, column=2).value = 'lstm'
    my_sheet.cell(row=model_counter, column=3).value = 'dense'
    my_sheet.cell(row=model_counter, column=4).value = 'epoch'
    my_sheet.cell(row=model_counter, column=5).value = 'batch_size'
    my_sheet.cell(row=model_counter, column=6).value = 'learning_rate'
    my_sheet.cell(row=model_counter, column=7).value = 'optimizer'
    my_sheet.cell(row=model_counter, column=8).value = 'weight_initialization'
    my_sheet.cell(row=model_counter, column=9).value = 'final_layer_activation'
    my_sheet.cell(row=model_counter, column=10).value = 'hidden_layer_activation'
    my_sheet.cell(row=model_counter, column=11).value = 'rmse'
    my_sheet.cell(row=model_counter, column=12).value = 'mape'
    my_wb.save("ModelsInfo.xlsx")
    model_counter += 1
else:
    my_wb = openpyxl.load_workbook('ModelsInfo.xlsx')
    my_sheet = my_wb.active
    model_counter = my_sheet.max_row + 1


def train_evaluate(ga_individual_solution):
    global model_counter
    global dense_units
    global window_size
    global lstm_units
    global lr
    global im
    global act
    global hact
    global bs
    global opt

    # Decode genetic algorithm solution to integer for window_size and num_units
    window_size_bits = BitArray(ga_individual_solution[0:7])
    lstm_units_bits = BitArray(ga_individual_solution[7:14])
    dense_units_bits = BitArray(ga_individual_solution[14:21])
    learning_rate_bits = BitArray(ga_individual_solution[21:28])
    optimizer_bits = BitArray(ga_individual_solution[28:35])
    init_mode_bits = BitArray(ga_individual_solution[35:42])
    activation_bits = BitArray(ga_individual_solution[42:49])
    hactivation_bits = BitArray(ga_individual_solution[49:56])
    batch_size_bits = BitArray(ga_individual_solution[56:63])
    window_size = window_size_bits.uint
    lstm_units = lstm_units_bits.uint
    dense_units = dense_units_bits.uint
    lr = learning_rate_bits.uint
    opt = optimizer_bits.uint
    im = init_mode_bits.uint
    act = activation_bits.uint
    hact = hactivation_bits.uint
    bs = batch_size_bits.uint

    # Return fitness score of 100 if window_size or num_unit is zero
    if window_size == 0 or lstm_units == 0 or dense_units == 0 or window_size < 6:
        return 1000,

    rmse = float(subprocess.check_call(["python deep-learning-train.py " + str(dense_units) + " " + str(lstm_units) + " " +
                                  str(window_size) + " " + str(lr) + " " + str(50) + " " + str(bs) + " " +
                                  str(act) + " " + str(hact) + " " + str(im) + " " + str(opt) + " " +
                                  str(model_counter)],
                                 shell=True))

    gc.collect()

    model_counter += 1

    return rmse,


def multi_parents(ind1, ind2):
    new_ind = list()

    new_ind.append(copy.deepcopy(ind1))
    new_ind.append(copy.deepcopy(ind2))

    new_inds = list()

    choices = [bool(random.getrandbits(1)) for i in range(0, 10)]

    new_inds.append([ind1[0: 7] if choices[0] else ind2[0: 7],
                     ind1[7: 14] if choices[1] else ind2[7: 14],
                     ind1[14: 21] if choices[2] else ind2[14: 21],
                     ind1[21: 28] if choices[3] else ind2[21: 28],
                     ind1[28: 35] if choices[4] else ind2[28: 35],
                     ind1[35: 42] if choices[5] else ind2[35: 42],
                     ind1[42: 49] if choices[6] else ind2[42: 49],
                     ind1[49: 56] if choices[7] else ind2[49: 56],
                     ind1[56: 63] if choices[8] else ind2[56: 63]])

    new_inds.append([ind1[0: 7] if not choices[0] else ind2[0: 7],
                     ind1[7: 14] if not choices[1] else ind2[7: 14],
                     ind1[14: 21] if not choices[2] else ind2[14: 21],
                     ind1[21: 28] if not choices[3] else ind2[21: 28],
                     ind1[28: 35] if not choices[4] else ind2[28: 35],
                     ind1[35: 42] if not choices[5] else ind2[35: 42],
                     ind1[42: 49] if not choices[6] else ind2[42: 49],
                     ind1[49: 56] if not choices[7] else ind2[49: 56],
                     ind1[56: 63] if not choices[8] else ind2[56: 63]])

    for i in range(0, 2):
        for j in range(0, 10):
            new_ind[i][0 + j * 7: 7 + j * 7] = new_inds[i][j][0: 7]

    return new_ind[0], new_ind[1]


population_size = 64
num_generations = 32
gene_length = 63

creator.create('FitnessMax', base.Fitness, weights=(-1.0,))
creator.create('Individual', list, fitness=creator.FitnessMax)

toolbox = base.Toolbox()
toolbox.register('binary', bernoulli.rvs, 0.5)
toolbox.register('individual', tools.initRepeat, creator.Individual, toolbox.binary, n=gene_length)
toolbox.register('population', tools.initRepeat, list, toolbox.individual)

toolbox.register('mate', multi_parents)
toolbox.register('mutate', tools.mutShuffleIndexes, indpb=0.6)
toolbox.register('select', tools.selRoulette)
toolbox.register('evaluate', train_evaluate)

population = toolbox.population(n=population_size)
r = algorithms.eaSimple(population, toolbox, cxpb=0.4, mutpb=0.1, ngen=num_generations, verbose=True)

best_individuals = tools.selBest(population, k=1)
best_window_size = None
best_lstm_units = None
best_dense_units = None

for bi in best_individuals:
    window_size_bits = BitArray(bi[0:7])
    lstm_units_bits = BitArray(bi[7:14])
    dense_units_bits = BitArray(bi[14:21])
    learning_rate_bits = BitArray(bi[21:28])
    optimizer_bits = BitArray(bi[28:35])
    init_mode_bits = BitArray(bi[35:42])
    activation_bits = BitArray(bi[42:49])
    hactivation_bits = BitArray(bi[49:56])
    batch_size_bits = BitArray(bi[56:63])
    best_window_size = window_size_bits.uint
    best_lstm_units = lstm_units_bits.uint
    best_dense_units = dense_units_bits.uint
    best_learning_rate = learning_rate_bits.uint
    best_optimizer = optimizer_bits.uint
    best_init_mode = init_mode_bits.uint
    best_activation = activation_bits.uint
    best_hactivation = hactivation_bits.uint
    best_batch_size = batch_size_bits.uint

    print('\nNum of best_window_size: ', best_window_size,
          ', Num of best_lstm_units: ', best_lstm_units,
          ', Num of best_dense_units: ', best_dense_units,
          ', Num of best_learning_rate: ', best_learning_rate,
          ', Num of best_optimizer: ', best_optimizer,
          ', Num of best_init_mode: ', best_init_mode,
          ', Num of best_activation: ', best_activation,
          ', Num of best_hactivation: ', best_hactivation,
          ', Num of best_batch_size: ', best_batch_size)
