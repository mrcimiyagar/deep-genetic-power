import keras

import copy
import os
import random

import joblib
import keras
import numpy as np
import pandas as pd
import tensorflow
import tensorflow_core
from keras.backend import tensorflow_backend
from keras.callbacks import ModelCheckpoint
from keras.losses import mean_absolute_percentage_error
from keras.metrics import Mean
from keras.optimizers import Adam, SGD, Adagrad, Nadam, Adamax, Adadelta
from keras.regularizers import l2
from keras.wrappers.scikit_learn import KerasRegressor
from numpy import ravel
from sklearn.linear_model import SGDClassifier, LinearRegression
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split as split, cross_validate, GridSearchCV

from keras.layers import LSTM, Input, Dense, Reshape, Conv1D, Lambda, concatenate, TimeDistributed, MaxPooling1D, \
    Flatten, Average, average, GlobalAveragePooling2D, GlobalAveragePooling1D
from keras.models import Model

from deap import base, creator, tools, algorithms
from scipy.stats import bernoulli
from bitstring import BitArray
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import MinMaxScaler, StandardScaler, LabelEncoder, Normalizer
from sklearn.svm import SVC, SVR, LinearSVC, LinearSVR
from tensorboard.plugins.hparams.api_pb2 import Session
from tensorflow_core.core.protobuf.config_pb2 import ConfigProto
from tensorflow_core.python.keras.optimizer_v2.rmsprop import RMSProp
import xlsxwriter
import gc
import openpyxl
from DataGenerators import DataGenerator

from tester import Tester

np.random.seed(1120)

parallel = 1


def prepare_dataset(window_size, batch_size):
    data = pd.read_excel('dataset1395.xlsx')
    df = pd.DataFrame(data,
                      columns=['H1', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'H10', 'H11', 'H12', 'H14', 'H14'
                          , 'H15', 'H16', 'H17', 'H18', 'H19', 'H20', 'H21', 'H22', 'H23', 'H24'])
    raw = list(np.array(df))

    data = pd.read_excel('dataset1396.xlsx')
    df = pd.DataFrame(data,
                      columns=['H1', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'H10', 'H11', 'H12', 'H14', 'H14'
                          , 'H15', 'H16', 'H17', 'H18', 'H19', 'H20', 'H21', 'H22', 'H23', 'H24'])
    raw.extend(list(np.array(df)))

    data = pd.read_excel('dataset1397.xlsx')
    df = pd.DataFrame(data,
                      columns=['H1', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'H10', 'H11', 'H12', 'H14', 'H14'
                          , 'H15', 'H16', 'H17', 'H18', 'H19', 'H20', 'H21', 'H22', 'H23', 'H24'])
    raw.extend(list(np.array(df)))

    data = pd.read_excel('dataset1398.xlsx')
    df = pd.DataFrame(data,
                      columns=['H1', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'H10', 'H11', 'H12', 'H14', 'H14'
                          , 'H15', 'H16', 'H17', 'H18', 'H19', 'H20', 'H21', 'H22', 'H23', 'H24'])
    raw.extend(list(np.array(df)))

    raw = np.array(raw)
    rowCount = raw.shape[0] * raw.shape[1]
    raw = raw.reshape(rowCount)
    raw = raw[0: rowCount - 24, ]
    timeSeries = list(raw)

    timeSeries.extend([0 for i in range(0, 24)])

    data = pd.read_excel('datasetW.xlsx')
    df = pd.DataFrame(data,
                      columns=['year', 'dayOfYear', 'month', 'dayOfWeek', 'hour', 'holiday', 'hourChange', 'specialDay',
                               'Temperature', 'Humidity', 'WeatherState'])

    additionalData = np.array(df)
    additionalData = additionalData[0:additionalData.shape[0] - 24, ]

    timeSeries = np.array(timeSeries)

    additionalData = np.array(additionalData)
    additionalData.resize((additionalData.shape[0], 12))
    temp = []
    temp[:] = additionalData[:, 0]
    additionalData[:, 0] = timeSeries[:]
    additionalData[:, 11] = temp[:]
    dataset = np.array(additionalData)
    dataset.resize((dataset.shape[0], 14))

    for i in range(0, dataset.shape[0]):
        if i < 24:
            dataset[i, 11] = dataset[i, 0]
            dataset[i, 12] = dataset[i, 0]
            dataset[i, 13] = dataset[i, 0]
            continue
        if i < 24 * 7:
            dataset[i, 11] = dataset[i - 24, 0]
            dataset[i, 12] = dataset[i, 0]
            dataset[i, 13] = dataset[i, 0]
            continue
        if i < 365 * 24:
            dataset[i, 11] = dataset[i - 24, 0]
            dataset[i, 12] = dataset[i - 24 * 7, 0]
            dataset[i, 13] = dataset[i, 0]
            continue
        dataset[i, 11] = dataset[i - 24, 0]
        dataset[i, 12] = dataset[i - 24 * 7, 0]
        dataset[i, 13] = dataset[i - 365 * 24, 0]

    dataset = dataset.astype('float32')

    x, y = list(), list()
    for i in range(window_size, dataset.shape[0] - 23):
        temp = list(dataset[i - window_size: i, ].flatten())
        for j in range(0, 24):
            for l in range(1, 14):
                temp.append(dataset[i + j, l])
        x.append(np.array(temp))
        y.append(list(dataset[i: i + 24, 0].flatten()))

    x, y = np.array(x), np.array(y)

    scaler_minmax = MinMaxScaler(feature_range=(0, 1))
    x = scaler_minmax.fit_transform(x)

    scaler = Normalizer()
    x = scaler.fit_transform(x)

    print(x[x.shape[0] - 1, ])

    x = x.reshape((x.shape[0], 1, window_size * 14 + 24 * 13))
    y = y.reshape((y.shape[0], 1, 24))

    if parallel > 1:
        x = np.concatenate([x for i in range(0, parallel)], axis=1)

    train_x, train_y, eval_x, eval_y, test_x, test_y = [], [], [], [], [], []

    train_x[:], train_y[:] = x[:200 * 72, ], y[:200 * 72, ]
    eval_x[:], eval_y[:] = x[200 * 72: 300 * 72, ], y[200 * 72: 300 * 72, ]
    test_x[:], test_y[:] = x[300 * 72:, ], y[300 * 72:, ]

    train_x, train_y, eval_x, eval_y, test_x, test_y = np.array(train_x), np.array(train_y), \
                                                       np.array(eval_x), np.array(eval_y), \
                                                       np.array(test_x), np.array(test_y)

    extra = test_x.shape[0] - (test_x.shape[0] // batch_size) * batch_size

    test_x = test_x[extra:, ]
    test_y = test_y[extra:, ]

    return train_x, train_y, eval_x, eval_y, test_x, test_y


dense_units = 0
lstm_units = 0
window_size = 0
lr = 0
ep = 0
bs = 0
act = 0
hact = 0
im = 0
opt = 0

if not os.path.exists('ModelsInfo.xlsx'):
    model_counter = 1
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    my_sheet.title = 'ModelsInfo'
    my_sheet.cell(row=model_counter, column=1).value = 'window'
    my_sheet.cell(row=model_counter, column=2).value = 'lstm'
    my_sheet.cell(row=model_counter, column=3).value = 'dense'
    my_sheet.cell(row=model_counter, column=4).value = 'epoch'
    my_sheet.cell(row=model_counter, column=5).value = 'batch_size'
    my_sheet.cell(row=model_counter, column=6).value = 'learning_rate'
    my_sheet.cell(row=model_counter, column=7).value = 'optimizer'
    my_sheet.cell(row=model_counter, column=8).value = 'weight_initialization'
    my_sheet.cell(row=model_counter, column=9).value = 'final_layer_activation'
    my_sheet.cell(row=model_counter, column=10).value = 'hidden_layer_activation'
    my_sheet.cell(row=model_counter, column=11).value = 'rmse'
    my_sheet.cell(row=model_counter, column=12).value = 'mape'
    my_wb.save("ModelsInfo.xlsx")
    model_counter += 1
else:
    my_wb = openpyxl.load_workbook('ModelsInfo.xlsx')
    my_sheet = my_wb.active
    model_counter = my_sheet.max_row + 1


def search(estimator_fn):
    global model_counter
    global kmodel_in

    learn_rates = [0.001, 0.01, 0.1, 0.2, 0.3]
    optimizers = [Adam, SGD, RMSProp, Adagrad, Adadelta, Adamax, Nadam]
    optimizer_names = ['Adam', 'SGD', 'RMSProp', 'Adagrad', 'Adadelta', 'Adamax', 'Nadam']
    init_modes = ['uniform', 'lecun_uniform', 'normal', 'zero', 'glorot_normal', 'glorot_uniform', 'he_normal',
                  'he_uniform']
    activations = ['softmax', 'softplus', 'softsign', 'relu', 'tanh', 'sigmoid', 'hard_sigmoid', 'linear']
    hidden_activations = ['softmax', 'softplus', 'softsign', 'relu', 'tanh', 'sigmoid', 'hard_sigmoid', 'linear']
    batch_sizes = [12, 24, 36, 72, 144]
    epochs = [25, 50, 75, 100]

    learn_rate = learn_rates[0] if 0 <= lr < 32 \
        else learn_rates[1] if 32 <= lr < 64 \
        else learn_rates[2] if 64 <= lr < 96 \
        else learn_rates[3]

    optimizer = optimizers[0] if 0 <= opt < 16 \
        else optimizers[1] if 16 <= opt < 32 \
        else optimizers[2] if 32 <= opt < 48 \
        else optimizers[3] if 48 <= opt < 64 \
        else optimizers[4] if 64 <= opt < 80 \
        else optimizers[5] if 80 <= opt < 96 \
        else optimizers[6]

    optimizer_name = optimizer_names[0] if 0 <= opt < 16 \
        else optimizer_names[1] if 16 <= opt < 32 \
        else optimizer_names[2] if 32 <= opt < 48 \
        else optimizer_names[3] if 48 <= opt < 64 \
        else optimizer_names[4] if 64 <= opt < 80 \
        else optimizer_names[5] if 80 <= opt < 96 \
        else optimizer_names[6]

    init_mode = init_modes[0] if 0 <= im < 16 \
        else init_modes[1] if 16 <= im < 32 \
        else init_modes[2] if 32 <= im < 48 \
        else init_modes[3] if 48 <= im < 64 \
        else init_modes[4] if 64 <= im < 80 \
        else init_modes[5] if 80 <= im < 96 \
        else init_modes[6] if 96 <= im < 112 \
        else init_modes[7]

    activation = activations[0] if 0 <= act < 16 \
        else activations[1] if 16 <= act < 32 \
        else activations[2] if 32 <= act < 48 \
        else activations[3] if 48 <= act < 64 \
        else activations[4] if 64 <= act < 80 \
        else activations[5] if 80 <= act < 96 \
        else activations[6] if 96 <= act < 112 \
        else activations[7]

    hidden_activation = hidden_activations[0] if 0 <= hact < 16 \
        else hidden_activations[1] if 16 <= hact < 32 \
        else hidden_activations[2] if 32 <= hact < 48 \
        else hidden_activations[3] if 48 <= hact < 64 \
        else hidden_activations[4] if 64 <= hact < 80 \
        else hidden_activations[5] if 80 <= hact < 96 \
        else hidden_activations[6] if 96 <= hact < 112 \
        else hidden_activations[7]

    batch_size = batch_sizes[0] if 0 <= bs < 24 \
        else batch_sizes[1] if 24 <= bs < 48 \
        else batch_sizes[2] if 48 <= bs < 72 \
        else batch_sizes[3] if 72 <= bs < 96 \
        else batch_sizes[4]

    epoch = epochs[0] if 0 <= ep < 32 \
        else epochs[1] if 32 <= ep < 64 \
        else epochs[2] if 64 <= ep < 96 \
        else epochs[3]

    my_wb = openpyxl.load_workbook('ModelsInfo.xlsx')
    my_sheet = my_wb.active
    for i in range(2, my_sheet.max_row):
        window_size_temp = my_sheet.cell(i, 1).value
        lstm_units_temp = my_sheet.cell(i, 2).value
        dense_units_temp = my_sheet.cell(i, 3).value
        epoch_temp = my_sheet.cell(i, 4).value
        batch_size_temp = my_sheet.cell(i, 5).value
        learn_rate_temp = my_sheet.cell(i, 6).value
        optimizer_name_temp = my_sheet.cell(i, 7).value
        init_mode_temp = my_sheet.cell(i, 8).value
        activation_temp = my_sheet.cell(i, 9).value
        hidden_activation_temp = my_sheet.cell(i, 10).value
        rmse_temp = float(my_sheet.cell(i, 11).value)

        if window_size == int(window_size_temp) and \
                lstm_units == int(lstm_units_temp) and \
                dense_units == int(dense_units_temp) and \
                epoch == int(epoch_temp) and \
                batch_size == int(batch_size_temp) and \
                learn_rate == float(learn_rate_temp) and \
                optimizer_name == optimizer_name_temp and \
                init_mode == init_mode_temp and \
                activation == activation_temp and \
                hidden_activation == hidden_activation_temp:
            print('skipping : ',
                  window_size, ' ',
                  lstm_units, ' ',
                  dense_units, ' ',
                  epoch, ' ',
                  batch_size, ' ',
                  learn_rate, ' ',
                  optimizer_name, ' ',
                  init_mode, ' ',
                  activation, ' ',
                  hidden_activation, ' ',
                  rmse_temp)

            return float(rmse_temp)

    print('\nNum of window_size: ', window_size,
          ', Num of lstm_units: ', lstm_units,
          ', Num of dense_units: ', dense_units,
          ', Num of learning_rate: ', learn_rate,
          ', Num of optimizer: ', optimizer_name,
          ', Num of init_mode: ', init_mode,
          ', Num of activation: ', activation,
          ', Num of hactivation: ', hidden_activation,
          ', Num of batch_size: ', batch_size,
          ', Num of epoch: ', epoch)

    X, y, test_x, test_y = prepare_dataset(window_size, batch_size=batch_size)

    kmodel_in = Input(batch_shape=(batch_size, X.shape[1], X.shape[2]))

    model = estimator_fn(learn_rate=learn_rate,
                         init_mode=init_mode,
                         optimizer=optimizer,
                         activation=activation,
                         hidden_activation=hidden_activation)

    model.fit(x=X, y=y, validation_data=[test_x, test_y],
              epochs=epoch, batch_size=batch_size, verbose=1)
    model.save(filepath='Model_' + str(model_counter))

    y_pred = model.predict(test_x, batch_size=batch_size)

    del model

    print('\nNum of window_size: ', window_size,
          ', Num of lstm_units: ', lstm_units,
          ', Num of dense_units: ', dense_units,
          ', Num of learning_rate: ', learn_rate,
          ', Num of optimizer: ', optimizer_name,
          ', Num of init_mode: ', init_mode,
          ', Num of activation: ', activation,
          ', Num of hactivation: ', hidden_activation,
          ', Num of batch_size: ', batch_size,
          ', Num of epoch: ', epoch)
    print(test_y[test_y.shape[0] - 1,])
    print(y_pred[y_pred.shape[0] - 1,])
    test_y = test_y.reshape((test_y.shape[0], test_y.shape[2]))
    y_pred = y_pred.reshape((y_pred.shape[0], y_pred.shape[2]))

    try:
        rmse = np.sqrt(mean_squared_error(test_y, y_pred))
        mape = np.mean(np.abs(test_y - y_pred) / test_y * 100)
        print('Validation RMSE: ', rmse, '\n')

        my_wb = openpyxl.load_workbook('ModelsInfo.xlsx')
        my_sheet = my_wb.active
        my_sheet.cell(row=model_counter, column=1).value = window_size
        my_sheet.cell(row=model_counter, column=2).value = lstm_units
        my_sheet.cell(row=model_counter, column=3).value = dense_units
        my_sheet.cell(row=model_counter, column=4).value = epoch
        my_sheet.cell(row=model_counter, column=5).value = batch_size
        my_sheet.cell(row=model_counter, column=6).value = learn_rate
        my_sheet.cell(row=model_counter, column=7).value = optimizer_name
        my_sheet.cell(row=model_counter, column=8).value = init_mode
        my_sheet.cell(row=model_counter, column=9).value = activation
        my_sheet.cell(row=model_counter, column=10).value = hidden_activation
        my_sheet.cell(row=model_counter, column=11).value = rmse
        my_sheet.cell(row=model_counter, column=12).value = mape
        my_wb.save("ModelsInfo.xlsx")

        model_counter += 1

        return rmse

    except:
        return 1000


def train_evaluate(ga_individual_solution):
    global kmodel_in
    global parallel
    global dense_units
    global window_size
    global lstm_units
    global lr
    global im
    global act
    global hact
    global ep
    global bs
    global opt

    # Decode genetic algorithm solution to integer for window_size and num_units
    window_size_bits = BitArray(ga_individual_solution[0:7])
    lstm_units_bits = BitArray(ga_individual_solution[7:14])
    dense_units_bits = BitArray(ga_individual_solution[14:21])
    learning_rate_bits = BitArray(ga_individual_solution[21:28])
    optimizer_bits = BitArray(ga_individual_solution[28:35])
    init_mode_bits = BitArray(ga_individual_solution[35:42])
    activation_bits = BitArray(ga_individual_solution[42:49])
    hactivation_bits = BitArray(ga_individual_solution[49:56])
    batch_size_bits = BitArray(ga_individual_solution[56:63])
    epoch_bits = BitArray(ga_individual_solution[63:70])
    window_size = window_size_bits.uint
    lstm_units = lstm_units_bits.uint
    dense_units = dense_units_bits.uint
    lr = learning_rate_bits.uint
    op = optimizer_bits.uint
    im = init_mode_bits.uint
    act = activation_bits.uint
    hact = hactivation_bits.uint
    bs = batch_size_bits.uint
    ep = epoch_bits.uint

    # Return fitness score of 100 if window_size or num_unit is zero
    if window_size == 0 or lstm_units == 0 or dense_units == 0 or window_size < 6:
        return 100,

    rmse = search(create_final_model)

    gc.collect()

    return rmse,


kmodel_in = {}


def create_model(index, last_layer_units, init_mode, activation, last_activation):
    global window_size
    global lstm_units
    global dense_units
    global kmodel_in

    kmodel1 = Lambda(lambda input_x: input_x[:, index: index + 1, 0: window_size * 14],
                     output_shape=(1, window_size * 14))(kmodel_in)
    kmodel2 = Lambda(lambda input_x: input_x[:, index: index + 1, window_size * 14: window_size * 14 + 24 * 13],
                     output_shape=(1, 24 * 13))(kmodel_in)

    kmodel3 = Reshape((window_size, 14, 1))(kmodel1)
    kmodel3 = Lambda(lambda input_x: (input_x[:, :, 0: 1, :]), output_shape=(window_size, 1))(kmodel3)
    kmodel3 = Reshape((1, window_size, 1))(kmodel3)

    kmodel1 = LSTM(lstm_units, kernel_initializer=init_mode, activation=activation)(kmodel1)
    kmodel1 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel1)
    kmodel1 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel1)
    kmodel1 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel1)
    kmodel1 = Reshape((1, dense_units))(kmodel1)

    kmodel2_main = Reshape((24, 13))(kmodel2)
    kmodel2 = Lambda(lambda input_x: input_x[:, :, :10], output_shape=(24, 10))(kmodel2_main)
    kmodel2 = Reshape((24, 10))(kmodel2)
    kmodel4 = Lambda(lambda input_x: input_x[:, :, 10:], output_shape=(24, 3))(kmodel2_main)
    kmodel4 = Reshape((24, 3))(kmodel4)

    kmodel2 = Flatten()(kmodel2)
    kmodel2 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel2)
    kmodel2 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel2)
    kmodel2 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel2)
    kmodel2 = Reshape((1, dense_units))(kmodel2)

    kmodel3 = TimeDistributed(Conv1D(filters=5, kernel_size=3, kernel_initializer=init_mode, activation=activation))(
        kmodel3)
    kmodel3 = TimeDistributed(MaxPooling1D(pool_size=2))(kmodel3)
    kmodel3 = TimeDistributed(Flatten())(kmodel3)
    kmodel3 = LSTM(lstm_units, stateful=True, return_sequences=True)(kmodel3)
    kmodel3 = LSTM(lstm_units // 5 if lstm_units > 5 else lstm_units, stateful=True)(kmodel3)
    kmodel3 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel3)
    kmodel3 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel3)
    kmodel3 = Reshape((1, dense_units))(kmodel3)

    kmodel4 = Flatten()(kmodel4)
    kmodel4 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel4)
    kmodel4 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel4)
    kmodel4 = Reshape((1, dense_units))(kmodel4)

    kmodel = keras.layers.concatenate([kmodel1, kmodel2, kmodel3, kmodel4], axis=1)
    kmodel = GlobalAveragePooling1D()(kmodel)
    kmodel = Dense(72, activation=activation, kernel_initializer=init_mode)(kmodel)
    kmodel = Dense(64, activation=activation, kernel_initializer=init_mode)(kmodel)
    kmodel = Dense(56, activation=activation, kernel_initializer=init_mode)(kmodel)
    kmodel = Dense(48, activation=activation, kernel_initializer=init_mode)(kmodel)
    kmodel = Dense(32, activation=activation, kernel_initializer=init_mode)(kmodel)
    kmodel = Dense(24, kernel_regularizer=l2(0.01), activation=last_activation, kernel_initializer=init_mode)(kmodel)

    return kmodel


def create_final_model(learn_rate=0.01,
                       init_mode='he_normal',
                       optimizer=Adam,
                       activation='linear',
                       hidden_activation='relu'):
    global lstm_units
    global dense_units

    kmodel = create_model(index=0, last_layer_units=dense_units, init_mode=init_mode, last_activation=activation,
                          activation=hidden_activation)
    kmodel = Model(inputs=[kmodel_in], outputs=[kmodel])
    kmodel.compile(optimizer=optimizer(learning_rate=learn_rate), loss='mse', metrics=['acc', 'mape'])
    return kmodel


window_size = 78
lstm_units = 21
dense_units = 72
batch_size = 24

X, y, eval_x, eval_y, test_x, test_y = prepare_dataset(window_size, batch_size=batch_size)

kmodel_in = Input(batch_shape=(24, X.shape[1], X.shape[2]))

model = create_final_model(learn_rate=0.001,
                           init_mode='he_normal',
                           optimizer=Adam,
                           activation='linear',
                           hidden_activation='relu')

if os.path.exists('PreTrained.h5'):
    model.load_weights('PreTrained.h5')
else:
    model.fit(X, y[:, 0, :], validation_data=[eval_x, eval_y[:, 0, :]], verbose=1, epochs=100, batch_size=batch_size)
    model.save('PreTrained.h5')


def extract_features(input_data, y):
    global batch_size

    y = y.reshape((y.shape[0] * 24))

    features = np.zeros(shape=(y.shape[0]))
    labels = np.zeros(shape=(y.shape[0]))

    generator = DataGenerator(x=input_data, y=y, batch_size=batch_size, shuffle=False)

    i = 0
    for input_batch, label_batch in generator:
        features_batch = model.predict(input_batch, batch_size=batch_size)
        features_batch = features_batch.reshape((features_batch.shape[0] * 24))
        features[i * 24 * batch_size: (i + 1) * 24 * batch_size] = features_batch
        labels[i * 24 * batch_size: (i + 1) * 24 * batch_size] = label_batch
        i += 1
        if i * batch_size * 24 >= y.shape[0]:
            break

    return features, labels


train_features, train_labels = extract_features(X, y)
validation_features, validation_labels = extract_features(eval_x, eval_y)
test_features, test_labels = extract_features(test_x, test_y)

svm_features = np.concatenate((train_features, validation_features), axis=0)
svm_labels = np.concatenate((train_labels, validation_labels), axis=0)

X_train, y_train = svm_features.reshape((svm_features.shape[0], 1)), svm_labels.reshape((svm_labels.shape[0]))
X_test, y_test = test_features.reshape((test_features.shape[0], 1)), test_labels.reshape((test_labels.shape[0]))

# svm = SVR(kernel='rbf', C=1, verbose=1)
# svm.fit(X_train, y_train)
# joblib.dump(svm, 'SvmModel_minmax')
svm = joblib.load('SvmModel')

tester = Tester('result_without_scaler')
tester.test_neuron(model, test_x, test_y, batch_size)
tester.test_svm(model, X_test, test_y, batch_size, svm, [])
