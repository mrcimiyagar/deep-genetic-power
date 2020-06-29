import copy
import os
import random

import keras
import numpy as np
import pandas as pd
from keras.optimizers import Adam, SGD, Adagrad, Nadam, Adamax, Adadelta
from sklearn.metrics import mean_squared_error

from keras.layers import LSTM, Input, Dense, Reshape, Conv1D, Lambda, TimeDistributed, MaxPooling1D, \
    Flatten, GlobalAveragePooling1D
from keras.models import Model

from deap import base, creator, tools, algorithms
from scipy.stats import bernoulli
from bitstring import BitArray
from sklearn.preprocessing import MinMaxScaler, Normalizer
from tensorflow_core.python.keras.optimizer_v2.rmsprop import RMSProp
import gc
import openpyxl

np.random.seed(1120)
random.seed(1120)

parallel = 1


def prepare_data_set(window_size, batch_size):

    data = pd.read_excel('datasetW.xlsx', sheet_name='input')
    df = pd.DataFrame(data,
                      columns=['year', 'dayOfYear', 'month', 'dayOfWeek', 'hour', 'holiday', 'hourChange', 'specialDay',
                               'Temperature', 'Humidity', 'WeatherState'])
    weather = np.array(df)
    data = pd.read_excel('datasetW.xlsx', sheet_name='output')
    df = pd.DataFrame(data, columns=['Load'])
    load = np.array(df)

    weather.resize((weather.shape[0], 12))
    temp = []
    temp[:] = weather[:, 0]
    weather[:, 0] = load[:, 0]
    weather[:, 11] = temp[:]
    dataset = np.array(weather)
    dataset.resize((dataset.shape[0], 14))

    for i in range(0, dataset.shape[0]):
        if i < 24:
            dataset[i, 11] = dataset[i, 0]
            dataset[i, 12] = dataset[i, 0]
            dataset[i, 13] = dataset[i, 0]
            continue
        if i < 24 * 7:
            dataset[i, 11] = dataset[i - 24, 0]
            dataset[i, 12] = dataset[i, 0]
            dataset[i, 13] = dataset[i, 0]
            continue
        if i < 365 * 24:
            dataset[i, 11] = dataset[i - 24, 0]
            dataset[i, 12] = dataset[i - 24 * 7, 0]
            dataset[i, 13] = dataset[i, 0]
            continue
        dataset[i, 11] = dataset[i - 24, 0]
        dataset[i, 12] = dataset[i - 24 * 7, 0]
        dataset[i, 13] = dataset[i - 365 * 24, 0]

    dataset = dataset.astype('float32')

    x, y = list(), list()
    for i in range(window_size, dataset.shape[0] - 23):
        temp = list(dataset[i - window_size: i, ].flatten())
        for j in range(0, 24):
            for l in range(1, 14):
                temp.append(dataset[i + j, l])
        x.append(np.array(temp))
        y.append(list(dataset[i: i + 24, 0].flatten()))

    x, y = np.array(x), np.array(y)

    scaler_minmax = MinMaxScaler(feature_range=(0, 1))
    x = scaler_minmax.fit_transform(x)

    normalizer = Normalizer()
    x = normalizer.fit_transform(x)

    x = x.reshape((x.shape[0], 1, window_size * 14 + 24 * 13))
    y = y.reshape((y.shape[0], 1, 24))

    trainX, trainY, evalX, evalY, testX, testY = [], [], [], [], [], []

    trainX[:], trainY[:] = x[:200 * 72, ], y[:200 * 72, ]
    evalX[:], evalY[:] = x[200 * 72: 300 * 72, ], y[200 * 72: 300 * 72, ]
    testX[:], testY[:] = x[300 * 72:, ], y[300 * 72:, ]

    trainX, trainY, evalX, evalY, testX, testY = np.array(trainX), np.array(trainY), \
                                                       np.array(evalX), np.array(evalY), \
                                                       np.array(testX), np.array(testY)

    extra = trainX.shape[0] - (trainX.shape[0] // batch_size) * batch_size
    trainX = trainX[extra:, ]
    trainY = trainY[extra:, ]

    extra = evalX.shape[0] - (evalX.shape[0] // batch_size) * batch_size
    evalX = evalX[extra:, ]
    evalY = evalY[extra:, ]

    extra = testX.shape[0] - (testX.shape[0] // batch_size) * batch_size
    testX = testX[extra:, ]
    testY = testY[extra:, ]

    return trainX, trainY, evalX, evalY, testX, testY


dense_units = 0
lstm_units = 0
window_size = 0
lr = 0
ep = 0
bs = 0
act = 0
hact = 0
im = 0
opt = 0

if not os.path.exists('ModelsInfo.xlsx'):
    model_counter = 1
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    my_sheet.title = 'ModelsInfo'
    my_sheet.cell(row=model_counter, column=1).value = 'window'
    my_sheet.cell(row=model_counter, column=2).value = 'lstm'
    my_sheet.cell(row=model_counter, column=3).value = 'dense'
    my_sheet.cell(row=model_counter, column=4).value = 'epoch'
    my_sheet.cell(row=model_counter, column=5).value = 'batch_size'
    my_sheet.cell(row=model_counter, column=6).value = 'learning_rate'
    my_sheet.cell(row=model_counter, column=7).value = 'optimizer'
    my_sheet.cell(row=model_counter, column=8).value = 'weight_initialization'
    my_sheet.cell(row=model_counter, column=9).value = 'final_layer_activation'
    my_sheet.cell(row=model_counter, column=10).value = 'hidden_layer_activation'
    my_sheet.cell(row=model_counter, column=11).value = 'rmse'
    my_sheet.cell(row=model_counter, column=12).value = 'mape'
    my_wb.save("ModelsInfo.xlsx")
    model_counter += 1
else:
    my_wb = openpyxl.load_workbook('ModelsInfo.xlsx')
    my_sheet = my_wb.active
    model_counter = my_sheet.max_row + 1


def search(estimator_fn):
    global model_counter
    global kmodel_in

    learn_rates = [0.001, 0.01, 0.1, 0.2, 0.3]
    optimizers = [Adam, SGD, RMSProp, Adagrad, Adadelta, Adamax, Nadam]
    optimizer_names = ['Adam', 'SGD', 'RMSProp', 'Adagrad', 'Adadelta', 'Adamax', 'Nadam']
    init_modes = ['uniform', 'lecun_uniform', 'normal', 'zero', 'glorot_normal', 'glorot_uniform', 'he_normal',
                  'he_uniform']
    activations = ['softmax', 'softplus', 'softsign', 'relu', 'tanh', 'sigmoid', 'hard_sigmoid', 'linear']
    hidden_activations = ['softmax', 'softplus', 'softsign', 'relu', 'tanh', 'sigmoid', 'hard_sigmoid', 'linear']
    batch_sizes = [12, 24, 36, 72, 144]
    epochs = [25, 50, 75, 100]

    learn_rate = learn_rates[0] if 0 <= lr < 32 \
        else learn_rates[1] if 32 <= lr < 64 \
        else learn_rates[2] if 64 <= lr < 96 \
        else learn_rates[3]

    optimizer = optimizers[0] if 0 <= opt < 16 \
        else optimizers[1] if 16 <= opt < 32 \
        else optimizers[2] if 32 <= opt < 48 \
        else optimizers[3] if 48 <= opt < 64 \
        else optimizers[4] if 64 <= opt < 80 \
        else optimizers[5] if 80 <= opt < 96 \
        else optimizers[6]

    optimizer_name = optimizer_names[0] if 0 <= opt < 16 \
        else optimizer_names[1] if 16 <= opt < 32 \
        else optimizer_names[2] if 32 <= opt < 48 \
        else optimizer_names[3] if 48 <= opt < 64 \
        else optimizer_names[4] if 64 <= opt < 80 \
        else optimizer_names[5] if 80 <= opt < 96 \
        else optimizer_names[6]

    init_mode = init_modes[0] if 0 <= im < 16 \
        else init_modes[1] if 16 <= im < 32 \
        else init_modes[2] if 32 <= im < 48 \
        else init_modes[3] if 48 <= im < 64 \
        else init_modes[4] if 64 <= im < 80 \
        else init_modes[5] if 80 <= im < 96 \
        else init_modes[6] if 96 <= im < 112 \
        else init_modes[7]

    activation = activations[0] if 0 <= act < 16 \
        else activations[1] if 16 <= act < 32 \
        else activations[2] if 32 <= act < 48 \
        else activations[3] if 48 <= act < 64 \
        else activations[4] if 64 <= act < 80 \
        else activations[5] if 80 <= act < 96 \
        else activations[6] if 96 <= act < 112 \
        else activations[7]

    hidden_activation = hidden_activations[0] if 0 <= hact < 16 \
        else hidden_activations[1] if 16 <= hact < 32 \
        else hidden_activations[2] if 32 <= hact < 48 \
        else hidden_activations[3] if 48 <= hact < 64 \
        else hidden_activations[4] if 64 <= hact < 80 \
        else hidden_activations[5] if 80 <= hact < 96 \
        else hidden_activations[6] if 96 <= hact < 112 \
        else hidden_activations[7]

    batch_size = batch_sizes[0] if 0 <= bs < 24 \
        else batch_sizes[1] if 24 <= bs < 48 \
        else batch_sizes[2] if 48 <= bs < 72 \
        else batch_sizes[3] if 72 <= bs < 96 \
        else batch_sizes[4]

    epoch = epochs[0] if 0 <= ep < 32 \
        else epochs[1] if 32 <= ep < 64 \
        else epochs[2] if 64 <= ep < 96 \
        else epochs[3]

    my_wb = openpyxl.load_workbook('ModelsInfo.xlsx')
    my_sheet = my_wb.active
    for i in range(2, my_sheet.max_row):
        window_size_temp = my_sheet.cell(i, 1).value
        lstm_units_temp = my_sheet.cell(i, 2).value
        dense_units_temp = my_sheet.cell(i, 3).value
        epoch_temp = my_sheet.cell(i, 4).value
        batch_size_temp = my_sheet.cell(i, 5).value
        learn_rate_temp = my_sheet.cell(i, 6).value
        optimizer_name_temp = my_sheet.cell(i, 7).value
        init_mode_temp = my_sheet.cell(i, 8).value
        activation_temp = my_sheet.cell(i, 9).value
        hidden_activation_temp = my_sheet.cell(i, 10).value
        rmse_temp = float(my_sheet.cell(i, 11).value)

        if window_size == int(window_size_temp) and \
                lstm_units == int(lstm_units_temp) and \
                dense_units == int(dense_units_temp) and \
                epoch == int(epoch_temp) and \
                batch_size == int(batch_size_temp) and \
                learn_rate == float(learn_rate_temp) and \
                optimizer_name == optimizer_name_temp and \
                init_mode == init_mode_temp and \
                activation == activation_temp and \
                hidden_activation == hidden_activation_temp:

            print('skipping : ',
                  window_size, ' ',
                  lstm_units, ' ',
                  dense_units, ' ',
                  epoch, ' ',
                  batch_size, ' ',
                  learn_rate, ' ',
                  optimizer_name, ' ',
                  init_mode, ' ',
                  activation, ' ',
                  hidden_activation, ' ',
                  rmse_temp)

            return float(rmse_temp)

    print('\nNum of window_size: ', window_size,
          ', Num of lstm_units: ', lstm_units,
          ', Num of dense_units: ', dense_units,
          ', Num of learning_rate: ', learn_rate,
          ', Num of optimizer: ', optimizer_name,
          ', Num of init_mode: ', init_mode,
          ', Num of activation: ', activation,
          ', Num of hactivation: ', hidden_activation,
          ', Num of batch_size: ', batch_size,
          ', Num of epoch: ', epoch)

    X, y, eval_x, eval_y, test_x, test_y = prepare_data_set(window_size, batch_size)

    kmodel_in = Input(batch_shape=(batch_size, X.shape[1], X.shape[2]))

    model = estimator_fn(learn_rate=learn_rate,
                         init_mode=init_mode,
                         optimizer=optimizer,
                         activation=activation,
                         hidden_activation=hidden_activation)

    model.fit(x=X, y=y[:, 0, :], validation_data=[test_x, test_y[:, 0, :]], shuffle=False,
              epochs=epoch, batch_size=batch_size, verbose=1, use_multiprocessing=True, workers=12)

    y_pred = model.predict(test_x, batch_size=batch_size)

    print('\nNum of window_size: ', window_size,
          ', Num of lstm_units: ', lstm_units,
          ', Num of dense_units: ', dense_units,
          ', Num of learning_rate: ', learn_rate,
          ', Num of optimizer: ', optimizer_name,
          ', Num of init_mode: ', init_mode,
          ', Num of activation: ', activation,
          ', Num of hactivation: ', hidden_activation,
          ', Num of batch_size: ', batch_size,
          ', Num of epoch: ', epoch)
    print(test_y[test_y.shape[0] - 1, ])
    print(y_pred[y_pred.shape[0] - 1, ])
    test_y = test_y.reshape((test_y.shape[0], 24))
    y_pred = y_pred.reshape((y_pred.shape[0], 24))

    rmse = np.sqrt(mean_squared_error(test_y, y_pred))
    mape = np.mean(np.abs(test_y - y_pred) / test_y * 100)
    print('Validation RMSE: ', rmse, '\n')
    model.save(filepath='Model_' + str(model_counter) + ".h5")

    my_wb = openpyxl.load_workbook('ModelsInfo.xlsx')
    my_sheet = my_wb.active
    my_sheet.cell(row=model_counter, column=1).value = window_size
    my_sheet.cell(row=model_counter, column=2).value = lstm_units
    my_sheet.cell(row=model_counter, column=3).value = dense_units
    my_sheet.cell(row=model_counter, column=4).value = epoch
    my_sheet.cell(row=model_counter, column=5).value = batch_size
    my_sheet.cell(row=model_counter, column=6).value = learn_rate
    my_sheet.cell(row=model_counter, column=7).value = optimizer_name
    my_sheet.cell(row=model_counter, column=8).value = init_mode
    my_sheet.cell(row=model_counter, column=9).value = activation
    my_sheet.cell(row=model_counter, column=10).value = hidden_activation
    my_sheet.cell(row=model_counter, column=11).value = rmse
    my_sheet.cell(row=model_counter, column=12).value = mape
    my_wb.save("ModelsInfo.xlsx")

    model_counter += 1

    del model
    del X
    del y
    del eval_x
    del eval_y
    del test_x
    del test_y

    return rmse


def train_evaluate(ga_individual_solution):
    global kmodel_in
    global parallel
    global dense_units
    global window_size
    global lstm_units
    global lr
    global im
    global act
    global hact
    global ep
    global bs
    global opt

    # Decode genetic algorithm solution to integer for window_size and num_units
    window_size_bits = BitArray(ga_individual_solution[0:7])
    lstm_units_bits = BitArray(ga_individual_solution[7:14])
    dense_units_bits = BitArray(ga_individual_solution[14:21])
    learning_rate_bits = BitArray(ga_individual_solution[21:28])
    optimizer_bits = BitArray(ga_individual_solution[28:35])
    init_mode_bits = BitArray(ga_individual_solution[35:42])
    activation_bits = BitArray(ga_individual_solution[42:49])
    hactivation_bits = BitArray(ga_individual_solution[49:56])
    batch_size_bits = BitArray(ga_individual_solution[56:63])
    epoch_bits = BitArray(ga_individual_solution[63:70])
    window_size = window_size_bits.uint
    lstm_units = lstm_units_bits.uint
    dense_units = dense_units_bits.uint
    lr = learning_rate_bits.uint
    op = optimizer_bits.uint
    im = init_mode_bits.uint
    act = activation_bits.uint
    hact = hactivation_bits.uint
    bs = batch_size_bits.uint
    ep = epoch_bits.uint

    # Return fitness score of 100 if window_size or num_unit is zero
    if window_size == 0 or lstm_units == 0 or dense_units == 0 or window_size < 6:
        return 1000,

    rmse = search(create_final_model)

    gc.collect()

    return rmse,


kmodel_in = {}


def create_model(index, last_layer_units, init_mode, activation, last_activation):
    global window_size
    global lstm_units
    global dense_units
    global kmodel_in

    kmodel1 = Lambda(lambda input_x: input_x[:, :, 0: window_size * 14],
                     output_shape=(1, window_size * 14))(kmodel_in)
    kmodel2 = Lambda(lambda input_x: input_x[:, :, window_size * 14: window_size * 14 + 24 * 13],
                     output_shape=(1, 24 * 13))(kmodel_in)

    kmodel3 = Reshape((window_size, 14, 1))(kmodel1)
    kmodel3 = Lambda(lambda input_x: (input_x[:, :, 0: 1, :]), output_shape=(window_size, 1))(kmodel3)
    kmodel3 = Reshape((1, window_size, 1))(kmodel3)

    kmodel1 = LSTM(lstm_units, kernel_initializer=init_mode, activation=activation)(kmodel1)
    kmodel1 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel1)
    kmodel1 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel1)
    kmodel1 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel1)
    kmodel1 = Reshape((1, dense_units))(kmodel1)

    kmodel2_main = Reshape((24, 13))(kmodel2)
    kmodel2 = Lambda(lambda input_x: input_x[:, :, :10], output_shape=(24, 10))(kmodel2_main)
    kmodel2 = Reshape((24, 10))(kmodel2)
    kmodel4 = Lambda(lambda input_x: input_x[:, :, 10:], output_shape=(24, 3))(kmodel2_main)
    kmodel4 = Reshape((24, 3))(kmodel4)

    kmodel2 = Flatten()(kmodel2)
    kmodel2 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel2)
    kmodel2 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel2)
    kmodel2 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel2)
    kmodel2 = Reshape((1, dense_units))(kmodel2)

    kmodel3 = TimeDistributed(Conv1D(filters=5, kernel_size=3, kernel_initializer=init_mode, activation=activation))(
        kmodel3)
    kmodel3 = TimeDistributed(MaxPooling1D(pool_size=2))(kmodel3)
    kmodel3 = TimeDistributed(Flatten())(kmodel3)
    kmodel3 = LSTM(lstm_units, stateful=True, return_sequences=True)(kmodel3)
    kmodel3 = LSTM(lstm_units // 5 if lstm_units > 5 else lstm_units, stateful=True)(kmodel3)
    kmodel3 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel3)
    kmodel3 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel3)
    kmodel3 = Reshape((1, dense_units))(kmodel3)

    kmodel4 = Flatten()(kmodel4)
    kmodel4 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel4)
    kmodel4 = Dense(dense_units, activation=activation, kernel_initializer=init_mode)(kmodel4)
    kmodel4 = Reshape((1, dense_units))(kmodel4)

    kmodel = keras.layers.concatenate([kmodel1, kmodel2, kmodel3, kmodel4], axis=1)
    kmodel = GlobalAveragePooling1D()(kmodel)
    kmodel = Dense(last_layer_units, activation=activation, kernel_initializer=init_mode)(kmodel)
    kmodel = Dense(24, activation=last_activation, kernel_initializer=init_mode)(kmodel)

    return kmodel


def create_final_model(learn_rate=0.01,
                       init_mode='uniform',
                       optimizer=Adam,
                       activation='linear',
                       hidden_activation='relu'):
    global lstm_units
    global dense_units

    kmodel = create_model(index=0, last_layer_units=dense_units, init_mode=init_mode, last_activation=activation,
                          activation=hidden_activation)
    kmodel = Model(inputs=[kmodel_in], outputs=[kmodel])
    kmodel.compile(optimizer=optimizer(learning_rate=learn_rate), loss='mse', metrics=['acc', 'mape'])
    return kmodel


def multi_parents(ind1, ind2):

    new_ind = list()

    new_ind.append(copy.deepcopy(ind1))
    new_ind.append(copy.deepcopy(ind2))

    new_inds = list()

    choices = [bool(random.getrandbits(1)) for i in range(0, 10)]

    new_inds.append([ind1[0: 7] if choices[0] else ind2[0: 7],
                     ind1[7: 14] if choices[1] else ind2[7: 14],
                     ind1[14: 21] if choices[2] else ind2[14: 21],
                     ind1[21: 28] if choices[3] else ind2[21: 28],
                     ind1[28: 35] if choices[4] else ind2[28: 35],
                     ind1[35: 42] if choices[5] else ind2[35: 42],
                     ind1[42: 49] if choices[6] else ind2[42: 49],
                     ind1[49: 56] if choices[7] else ind2[49: 56],
                     ind1[56: 63] if choices[8] else ind2[56: 63],
                     ind1[63: 70] if choices[9] else ind2[63: 70]])

    new_inds.append([ind1[0: 7] if not choices[0] else ind2[0: 7],
                     ind1[7: 14] if not choices[1] else ind2[7: 14],
                     ind1[14: 21] if not choices[2] else ind2[14: 21],
                     ind1[21: 28] if not choices[3] else ind2[21: 28],
                     ind1[28: 35] if not choices[4] else ind2[28: 35],
                     ind1[35: 42] if not choices[5] else ind2[35: 42],
                     ind1[42: 49] if not choices[6] else ind2[42: 49],
                     ind1[49: 56] if not choices[7] else ind2[49: 56],
                     ind1[56: 63] if not choices[8] else ind2[56: 63],
                     ind1[63: 70] if not choices[9] else ind2[63: 70]])

    for i in range(0, 2):
        for j in range(0, 10):
            new_ind[i][0 + j * 7: 7 + j * 7] = new_inds[i][j][0: 7]

    return new_ind[0], new_ind[1]

population_size = 64
num_generations = 64
gene_length = 70

# As we are trying to minimize the RMSE score, that's why using -1.0.
# In case, when you want to maximize accuracy for instance, use 1.0
creator.create('FitnessMax', base.Fitness, weights=(-1.0,))
creator.create('Individual', list, fitness=creator.FitnessMax)

toolbox = base.Toolbox()
toolbox.register('binary', bernoulli.rvs, 0.5)
toolbox.register('individual', tools.initRepeat, creator.Individual, toolbox.binary, n=gene_length)
toolbox.register('population', tools.initRepeat, list, toolbox.individual)

toolbox.register('mate', multi_parents)
toolbox.register('mutate', tools.mutShuffleIndexes, indpb=0.6)
toolbox.register('select', tools.selRoulette)
toolbox.register('evaluate', train_evaluate)

population = toolbox.population(n=population_size)
r = algorithms.eaSimple(population, toolbox, cxpb=0.4, mutpb=0.1, ngen=num_generations, verbose=True)

best_individuals = tools.selBest(population, k=1)
best_window_size = None
best_lstm_units = None
best_dense_units = None

for bi in best_individuals:
    window_size_bits = BitArray(bi[0:7])
    lstm_units_bits = BitArray(bi[7:14])
    dense_units_bits = BitArray(bi[14:21])
    learning_rate_bits = BitArray(bi[21:28])
    optimizer_bits = BitArray(bi[28:35])
    init_mode_bits = BitArray(bi[35:42])
    activation_bits = BitArray(bi[42:49])
    hactivation_bits = BitArray(bi[49:56])
    batch_size_bits = BitArray(bi[56:63])
    epoch_bits = BitArray(bi[63:70])
    best_window_size = window_size_bits.uint
    best_lstm_units = lstm_units_bits.uint
    best_dense_units = dense_units_bits.uint
    best_learning_rate = learning_rate_bits.uint
    best_optimizer = optimizer_bits.uint
    best_init_mode = init_mode_bits.uint
    best_activation = activation_bits.uint
    best_hactivation = hactivation_bits.uint
    best_batch_size = batch_size_bits.uint
    best_epoch = epoch_bits.uint

    print('\nNum of best_window_size: ', best_window_size,
          ', Num of best_lstm_units: ', best_lstm_units,
          ', Num of best_dense_units: ', best_dense_units,
          ', Num of best_learning_rate: ', best_learning_rate,
          ', Num of best_optimizer: ', best_optimizer,
          ', Num of best_init_mode: ', best_init_mode,
          ', Num of best_activation: ', best_activation,
          ', Num of best_hactivation: ', best_hactivation,
          ', Num of best_batch_size: ', best_batch_size,
          ', Num of best_epoch: ', best_epoch)
